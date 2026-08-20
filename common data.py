# -*- coding: utf-8 -*-
"""
데이터 로딩 모듈
================
원본 Excel(data/*.xlsx)을 직접 읽어 정제한다. 별도의 전처리 스크립트나
parquet 캐시 파일 없이, Streamlit의 @st.cache_data로 세션 내에서만 캐싱한다.

원자료 시트는 2단 헤더 구조를 가진다.
    행0 : (거의 공백) "(단위: 백만원)" 같은 단위 표기
    행1 : 실제 컬럼명 (기관명, 기관유형, 주무부처, [구분], [항목], 2021년~2026년, 상위기관)
    행2~: 실제 데이터
"""

import os
import re
import unicodedata

import numpy as np
import pandas as pd
import streamlit as st

from utils.constants import RAW_DIR

YEAR_RE = re.compile(r"^(20\d{2})년")

# key -> (파일명, 시트명, 연도 외 식별컬럼)
FILES = {
    "finance": ("수입지출현황.xlsx", "수입지출현황(고유사업)", ["항목"]),
    "tax": ("법인세정보.xlsx", "법인세", ["항목"]),
    "employees": ("임직원수현황.xlsx", "1. 임직원 수", ["항목"]),
    "compensation": ("직원평균보수현황.xlsx", "1. 직원평균보수", ["구분", "항목"]),
    "starting_pay": ("직원평균보수현황.xlsx", "2. 신입사원초임", ["항목"]),
    "executive_pay": ("임원연봉.xlsx", "임원연봉", ["구분", "항목"]),
    "recruitment": ("신규채용현황.xlsx", "1. 신규채용현황", ["항목"]),
    "welfare": ("복리후생비.xlsx", "1. 예산상 복리후생비", ["구분", "항목"]),
    "business_expense": ("기관장업무추진비.xlsx", "기관장업무추진비", ["항목"]),
    "other_welfare": ("그밖의_복리후생제도_등의_운영현황.xlsx", "1-2. 휴직급여지급현황", ["구분", "항목"]),
    "parental_leave": ("일가정_양립_지원제도_운영현황.xlsx", "1. 일가정-육아휴직사용자수", ["구분"]),
}


def _resolve_path(fname: str) -> str:
    """
    폴더 안에서 파일을 찾는다. 한글 파일명은 macOS(NFD)와 Linux/Windows(NFC) 간
    유니코드 정규화 방식이 달라 화면상 이름이 같아도 실제 바이트가 다를 수 있으므로,
    정확히 일치하지 않으면 정규화 후 비교하여 찾는다.
    """
    direct = os.path.join(RAW_DIR, fname)
    if os.path.exists(direct):
        return direct

    if not os.path.isdir(RAW_DIR):
        return direct

    target = unicodedata.normalize("NFC", fname)
    for f in os.listdir(RAW_DIR):
        if unicodedata.normalize("NFC", f) == target:
            return os.path.join(RAW_DIR, f)

    return direct


def normalize_institution_name(name):
    """기관명 표기 차이(공백 등)를 정규화한다."""
    if pd.isna(name):
        return name
    return re.sub(r"\s+", "", str(name)).strip()


def _find_header_row(raw_df, key_col="기관명", search_rows=5):
    for i in range(min(search_rows, len(raw_df))):
        if key_col in raw_df.iloc[i].astype(str).tolist():
            return i
    raise ValueError(f"'{key_col}' 헤더 행을 찾을 수 없습니다.")


def _clean_sheet(raw_df):
    """header=None으로 읽은 raw DataFrame에서 실제 헤더를 찾아 정리한다."""
    hdr_idx = _find_header_row(raw_df)
    header = [c.strip() if isinstance(c, str) else c for c in raw_df.iloc[hdr_idx]]
    df = raw_df.iloc[hdr_idx + 1:].copy()
    df.columns = header
    df = df[df["기관명"].notna()].reset_index(drop=True)

    year_cols = {c for c in df.columns if isinstance(c, str) and YEAR_RE.match(c)}
    for col in df.columns:
        if col in year_cols:
            continue
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)
    return df


@st.cache_data(show_spinner="원자료를 읽는 중...")
def _load_long(key: str) -> pd.DataFrame:
    """지정된 데이터셋을 raw Excel에서 읽어 long(연도) 포맷으로 반환한다."""
    fname, sheet, extra_ids = FILES[key]
    path = _resolve_path(fname)
    if not os.path.exists(path):
        return pd.DataFrame()

    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    df = _clean_sheet(raw)

    year_cols = [c for c in df.columns if isinstance(c, str) and YEAR_RE.match(c)]
    id_cols = [c for c in ["기관명", "기관유형", "주무부처"] if c in df.columns]
    id_cols += [c for c in extra_ids if c in df.columns]

    long_df = df.melt(id_vars=id_cols, value_vars=year_cols, var_name="연도", value_name="값")
    long_df["연도"] = long_df["연도"].str.extract(r"(20\d{2})").astype(int)
    long_df["값"] = pd.to_numeric(long_df["값"], errors="coerce")
    long_df["기관명"] = long_df["기관명"].apply(normalize_institution_name)
    return long_df


def _pivot_items(long_df, item_map, extra_filter=None, item_col="항목",
                  id_cols=("기관명", "기관유형", "주무부처", "연도")):
    """long_df에서 특정 항목만 뽑아 wide(item_map 값이 컬럼명)로 변환한다."""
    df = long_df.copy()
    if extra_filter:
        for k, v in extra_filter.items():
            if k in df.columns:
                df = df[df[k] == v]
    df = df[df[item_col].isin(item_map.keys())].copy()
    df[item_col] = df[item_col].map(item_map)
    id_cols = [c for c in id_cols if c in df.columns]
    wide = df.pivot_table(index=id_cols, columns=item_col, values="값", aggfunc="first").reset_index()
    wide.columns.name = None
    return wide


def _safe_ratio(numerator, denominator):
    """분모가 0이거나 결측이면 NaN을 반환하는 안전한 비율 계산."""
    num = pd.to_numeric(numerator, errors="coerce")
    den = pd.to_numeric(denominator, errors="coerce")
    ratio = num / den.replace(0, np.nan)
    return ratio.replace([np.inf, -np.inf], np.nan)


def _load_daycare() -> pd.DataFrame:
    """
    직장어린이집운영비 시트는 다른 시트와 달리 연도가 wide 컬럼이 아니라
    레코드별 값으로 존재하므로 별도 로직으로 처리한다.
    기관명+연도 기준으로 금액(daycare_expense)과 수혜인원(daycare_beneficiaries)을 합산한다.
    각 레코드가 세부 지원항목별로 나뉘어 있어 합산 시 항목 간 중복 가능성이 있을 수 있다.
    """
    fname = "일가정_양립_지원제도_운영현황.xlsx"
    sheet = "7. 일가정-직장어린이집운영비"
    path = _resolve_path(fname)
    if not os.path.exists(path):
        return pd.DataFrame()

    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    df = _clean_sheet(raw)
    if "연도" not in df.columns:
        return pd.DataFrame()

    df["연도"] = df["연도"].astype(str).str.extract(r"(20\d{2})").astype(float).astype("Int64")
    df["금액"] = pd.to_numeric(df["금액"], errors="coerce")
    df["수혜인원"] = pd.to_numeric(df["수혜인원"], errors="coerce")
    df["기관명"] = df["기관명"].apply(normalize_institution_name)

    grouped = df.groupby(["기관명", "기관유형", "주무부처", "연도"], as_index=False).agg(
        daycare_expense=("금액", "sum"), daycare_beneficiaries=("수혜인원", "sum")
    )
    grouped["연도"] = grouped["연도"].astype(int)
    return grouped


@st.cache_data(show_spinner="기관-연도 통합 분석 패널을 구성하는 중...")
def build_analysis_panel() -> pd.DataFrame:
    """
    여러 데이터셋의 핵심 변수를 기관명+연도 기준으로 통합한 분석용 데이터프레임.
    값이 없는 경우 절대 0으로 대체하지 않고 NaN을 유지한다 (outer merge).
    """
    finance = _load_long("finance")
    tax = _load_long("tax")
    employees = _load_long("employees")
    compensation = _load_long("compensation")
    starting_pay = _load_long("starting_pay")
    exec_pay = _load_long("executive_pay")
    recruitment = _load_long("recruitment")
    welfare = _load_long("welfare")
    biz_expense = _load_long("business_expense")
    parental = _load_long("parental_leave")

    if finance.empty:
        return pd.DataFrame()

    # --- 재무 ---
    panel = _pivot_items(finance, {
        "수입 > 수입합계": "total_revenue",
        "수입 > 정부지원수입 > 소계": "gov_support_revenue",
        "수입 > 기타사업수입": "business_revenue",
        "지출 > 지출합계": "total_expense",
        "지출 > 인건비": "labor_cost",
        "지출 > 경상운영비": "operating_cost",
        "지출 > 사업비": "business_cost",
    })

    # --- 법인세 (세부항목) ---
    tax_wide = _pivot_items(tax, {
        "과세표준": "taxable_income",
        "법인세 산출세액": "corporate_tax_calculated",
        "세액공제": "tax_credit",
        "가산세": "additional_tax",
        "결정세액": "corporate_tax_final",
    })

    # --- 임직원 (정원/현원, 여성) ---
    employees_wide = _pivot_items(employees, {
        "임직원 총계(A+B+C)": "total_workforce",
        "여성 현원-합계": "female_workforce",
    })
    authorized_items = ["정규직-일반정규직-정원-계(B)", "정규직-무기계약직-정원-계(C)"]
    current_items = ["정규직-일반정규직-현원-계", "정규직-무기계약직-현원-계"]
    auth_df = employees[employees["항목"].isin(authorized_items)]
    auth_wide = auth_df.pivot_table(index=["기관명", "기관유형", "주무부처", "연도"], columns="항목", values="값", aggfunc="first").reset_index()
    auth_wide.columns.name = None
    auth_present = [c for c in authorized_items if c in auth_wide.columns]
    auth_wide["total_authorized"] = auth_wide[auth_present].sum(axis=1, skipna=True) if auth_present else np.nan
    auth_wide = auth_wide[["기관명", "기관유형", "주무부처", "연도", "total_authorized"]]

    cur_df = employees[employees["항목"].isin(current_items)]
    cur_wide = cur_df.pivot_table(index=["기관명", "기관유형", "주무부처", "연도"], columns="항목", values="값", aggfunc="first").reset_index()
    cur_wide.columns.name = None
    cur_present = [c for c in current_items if c in cur_wide.columns]
    cur_wide["_current_regular"] = cur_wide[cur_present].sum(axis=1, skipna=True) if cur_present else np.nan
    cur_wide = cur_wide[["기관명", "기관유형", "주무부처", "연도", "_current_regular"]]

    nonreg_items = ["비정규직-기간제-계", "비정규직-기타", "비정규직-소속외 인력-계"]
    nonreg = employees[employees["항목"].isin(nonreg_items)]
    nonreg_wide = nonreg.pivot_table(index=["기관명", "기관유형", "주무부처", "연도"], columns="항목", values="값", aggfunc="first").reset_index()
    nonreg_wide.columns.name = None
    nonreg_present = [c for c in nonreg_items if c in nonreg_wide.columns]
    nonreg_wide["nonregular_workforce"] = nonreg_wide[nonreg_present].sum(axis=1, skipna=True) if nonreg_present else np.nan
    nonreg_wide = nonreg_wide[["기관명", "기관유형", "주무부처", "연도", "nonregular_workforce"]]

    # --- 신규채용 ---
    recruitment_wide = _pivot_items(recruitment, {
        "일반정규직총신규채용": "total_new_hires",
        "여성": "female_hires",
        "청년": "youth_hires",
        "장애인": "disabled_hires",
    })

    # --- 직원보수 (정규직(일반정규직) 기준) ---
    compensation_wide = _pivot_items(
        compensation,
        {
            "1인당 평균보수액": "employee_avg_pay", "평균근속연수(개월)": "avg_tenure_months",
            "기본급": "base_pay", "고정수당": "fixed_allowance",
            "성과상여금": "performance_pay", "(경영평가 성과급)": "management_eval_bonus",
        },
        extra_filter={"구분": "정규직(일반정규직)"},
    )
    starting_pay_wide = _pivot_items(starting_pay, {"합계": "starting_pay"})

    # --- 임원보수 ---
    exec_pay_wide = _pivot_items(exec_pay, {"합계": "executive_total_pay"}, extra_filter={"구분": "상임기관장"})

    # --- 업무추진비 ---
    biz_expense_wide = _pivot_items(biz_expense, {"업무추진비 집행금액": "executive_expense"})

    # --- 복리후생 ---
    welfare_total = welfare[welfare["항목"].astype(str).str.contains("총계", na=False)]
    welfare_item_name = welfare_total["항목"].iloc[0] if not welfare_total.empty else None
    welfare_wide = _pivot_items(welfare, {welfare_item_name: "total_welfare_expense"}) if welfare_item_name else pd.DataFrame()

    # --- 직장어린이집 ---
    daycare_wide = _load_daycare()

    # --- 육아휴직 (일가정 양립) ---
    parental_wide = _pivot_items(
        parental,
        {
            "전체 사용자 수": "parental_leave_total",
            "남성 사용자 수": "parental_leave_male",
            "여성 사용자 수": "parental_leave_female",
            "남성 육아휴직 사용률": "male_parental_leave_ratio_pct",
        },
        item_col="구분",
    )

    for wide_df in [tax_wide, employees_wide, auth_wide, cur_wide, nonreg_wide,
                    recruitment_wide, compensation_wide, starting_pay_wide,
                    exec_pay_wide, biz_expense_wide, welfare_wide, parental_wide, daycare_wide]:
        if wide_df is None or wide_df.empty:
            continue
        merge_cols = [c for c in ["기관명", "기관유형", "주무부처", "연도"] if c in wide_df.columns]
        panel = panel.merge(wide_df, on=merge_cols, how="outer")

    # --- 파생변수 (분자/분모 결측 또는 0이면 NaN) ---
    if {"_current_regular", "total_authorized"}.issubset(panel.columns):
        panel["fill_rate_pct"] = _safe_ratio(panel["_current_regular"], panel["total_authorized"]) * 100
        panel.drop(columns=["_current_regular"], inplace=True)

    if {"female_workforce", "total_workforce"}.issubset(panel.columns):
        panel["female_ratio_pct"] = _safe_ratio(panel["female_workforce"], panel["total_workforce"]) * 100

    if {"total_new_hires", "total_workforce"}.issubset(panel.columns):
        panel["new_hire_rate_pct"] = _safe_ratio(panel["total_new_hires"], panel["total_workforce"]) * 100
    if {"youth_hires", "total_new_hires"}.issubset(panel.columns):
        panel["youth_hire_ratio_pct"] = _safe_ratio(panel["youth_hires"], panel["total_new_hires"]) * 100
    if {"female_hires", "total_new_hires"}.issubset(panel.columns):
        panel["female_hire_ratio_pct"] = _safe_ratio(panel["female_hires"], panel["total_new_hires"]) * 100
    if {"disabled_hires", "total_new_hires"}.issubset(panel.columns):
        panel["disabled_hire_ratio_pct"] = _safe_ratio(panel["disabled_hires"], panel["total_new_hires"]) * 100

    if {"executive_total_pay", "employee_avg_pay"}.issubset(panel.columns):
        panel["executive_pay_multiple"] = _safe_ratio(panel["executive_total_pay"], panel["employee_avg_pay"])
    if {"executive_expense", "total_workforce"}.issubset(panel.columns):
        panel["executive_expense_per_capita"] = _safe_ratio(panel["executive_expense"], panel["total_workforce"])

    if {"total_welfare_expense", "total_workforce"}.issubset(panel.columns):
        panel["welfare_per_capita"] = _safe_ratio(panel["total_welfare_expense"], panel["total_workforce"])

    if {"parental_leave_total", "female_workforce"}.issubset(panel.columns):
        panel["parental_leave_rate_pct"] = _safe_ratio(panel["parental_leave_total"], panel["female_workforce"]) * 100

    if {"total_revenue", "total_expense"}.issubset(panel.columns):
        panel["balance"] = panel["total_revenue"] - panel["total_expense"]
    if {"gov_support_revenue", "total_revenue"}.issubset(panel.columns):
        panel["gov_dependency_pct"] = _safe_ratio(panel["gov_support_revenue"], panel["total_revenue"]) * 100
    if {"business_revenue"}.issubset(panel.columns):
        panel["own_revenue_conservative"] = panel["business_revenue"]
    if {"total_revenue", "gov_support_revenue"}.issubset(panel.columns):
        panel["own_revenue_broad"] = panel["total_revenue"] - panel["gov_support_revenue"]
    if {"total_revenue", "total_workforce"}.issubset(panel.columns):
        panel["revenue_per_employee"] = _safe_ratio(panel["total_revenue"], panel["total_workforce"])
    if {"business_revenue", "total_workforce"}.issubset(panel.columns):
        panel["business_revenue_per_employee"] = _safe_ratio(panel["business_revenue"], panel["total_workforce"])
    if {"labor_cost", "total_expense"}.issubset(panel.columns):
        panel["labor_cost_ratio_pct"] = _safe_ratio(panel["labor_cost"], panel["total_expense"]) * 100

    # --- 법인세 파생변수 (임의로 "실효세율"이라 명명하지 않음) ---
    if {"tax_credit", "corporate_tax_calculated"}.issubset(panel.columns):
        panel["tax_credit_ratio"] = _safe_ratio(panel["tax_credit"], panel["corporate_tax_calculated"]) * 100
    if {"corporate_tax_final", "taxable_income"}.issubset(panel.columns):
        panel["tax_burden_ratio"] = _safe_ratio(panel["corporate_tax_final"], panel["taxable_income"]) * 100

    return panel


def load_dataset(key: str) -> pd.DataFrame:
    """
    데이터셋을 불러온다.
    key가 'panel'이면 기관-연도 통합 분석 패널을, 그 외에는 해당 데이터셋의
    long(연도) 포맷을 반환한다. 원본 파일이 없거나 읽기에 실패하면
    구체적인 진단 정보와 함께 안내 메시지를 표시한다.
    """
    if key == "panel":
        df = build_analysis_panel()
    elif key in FILES:
        df = _load_long(key)
    else:
        raise ValueError(f"알 수 없는 데이터셋: {key}")

    if df.empty:
        if key == "panel":
            check_keys = ["finance", "tax", "employees", "compensation",
                          "executive_pay", "recruitment", "welfare", "business_expense"]
            target_key = next((k for k in check_keys if _load_long(k).empty), "finance")
        else:
            target_key = key

        info = _diagnose(target_key)
        st.error(f"'{target_key}' 데이터셋을 불러오지 못했습니다. 아래 진단 정보를 확인해주세요.")
        st.json(info)

    return df


def _diagnose(key: str) -> dict:
    """key에 해당하는 데이터셋이 비어있는 이유를 진단한다 (오류를 던지지 않음)."""
    fname, sheet, extra_ids = FILES[key]
    path = _resolve_path(fname)
    info = {"데이터셋": key, "파일명": fname, "찾은 경로": path, "파일 존재": os.path.exists(path)}

    if not info["파일 존재"]:
        return info

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        info["실제 시트 목록"] = wb.sheetnames
        info["찾는 시트명"] = sheet
        info["시트 일치 여부"] = sheet in wb.sheetnames
    except Exception as e:
        info["워크북 열기 오류"] = str(e)
        return info

    if not info["시트 일치 여부"]:
        return info

    try:
        raw = pd.read_excel(path, sheet_name=sheet, header=None)
        info["원본 시트 shape"] = raw.shape
        hdr_idx = _find_header_row(raw)
        info["헤더 행 위치"] = hdr_idx
        info["헤더 행 내용"] = raw.iloc[hdr_idx].tolist()
    except Exception as e:
        info["시트 읽기 오류"] = str(e)

    return info


def raw_files_exist() -> bool:
    return all(os.path.exists(_resolve_path(fname)) for fname, _, _ in FILES.values())
